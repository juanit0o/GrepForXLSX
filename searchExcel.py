import os, sys
import openpyxl

def searchIn(path, word):
    
    print("\n")

    #List of tuples of findings (FileName, PositionTable)
    l = []
    #Files in the path given as argument
    d = os.listdir(path)

    for file in d:
        #name of each file in the path
        filename = str(path) + "/" + str(file)
        #print("Finding %s in %s" % (word, file))
        print("Looking for " + word  + " in " + file)
        

        #Check if it is an excel file
        if filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(filename=filename)
            ws = wb.active

            for row in ws.rows:
                for cell in row:
                    if cell.value and str(word) in cell.value:
                        l.append((file, cell))
            
    
    #If it was found in all files combined 
    print("\n")
    if l:
        print("Word "+ word + " found "+ str(len(l)) + " times in:\n" )

        for fn, cell in l:
            #print("FileName: " +  fn + " located in =>  row: " + str(cell.row) + ",  column: "+ str(cell.column))
            print("File: '" + fn + "' Cell =>  row: " + str(cell.row) + ",  column: "+ str(cell.column))
            
            
            filename = str(path) + "/" + str(fn)
            #print("Finding %s in %s" % (word, file))
            
            #Check if it is an excel file
            if filename.endswith(".xlsx"):
                #Close the last workbook or it will always use the last file
                wb.close()
                wb = openpyxl.load_workbook(filename=filename)
                ws = wb.active

                line = " | "
                i = 1
                for row in ws.rows:
                    if(i > ws.max_column):
                            break
                    if(i == cell.row):
                        for cell in row:
                            
                            #Only add the row that contains the value
                            if(cell.value != None):
                                line += str(cell.value) + " | "
                            i += 1         
                    i+= 1
                
                print(line)
                print("=============================")        
    else:
        print("Word " + word + " not found :(")


if __name__ == "__main__":
    try:
        #Search in Path for word
        searchIn(sys.argv[1], sys.argv[2])
    except IndexError:
        print("\tExecute: python searchpy <path> <word>")
        print("\tEg: python searchExcel.py /pathtoFolder/ value1\n")
        print("\tIf you want to search from the current folder do")
        print("\tpython searchExcel.py './Excel Files' valor2")
